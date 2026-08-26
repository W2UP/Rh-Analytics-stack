from fastapi import FastAPI, BackgroundTasks, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import requests
from datetime import datetime, timedelta 
import calendar 
import urllib3
import concurrent.futures
import time 


import sqlite3
import json
import pandas as pd
import unicodedata
import io
import re
import openpyxl
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

app = FastAPI(title="RH Analytics ERP - Folha Edition")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], 
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

NOTION_TOKEN = "ntn_b86757525552BOl8M9h0NsNEntm9aZiTiVxCiwkNa8Kdue"

DB_COLABORADORES = "3ba2051d7a6b8005b98af9fab951026d"
DB_DESLIGAMENTOS = "3ab2051d7a6b808bb0d8fe4199cd5ed9"
DB_ATESTADOS = "ec42d14b9f4243a4ae42a4e704396b1c"
DB_ADVERTENCIAS = "3982051d7a6b8012b894f48c52cdab79"
DB_FREQUENCIA = "39d2051d7a6b805cac20cb52b5c0b476"
DB_DESEMPENHO = "14378c804c6643a3864e72d7947c822f"
DB_ARMARIOS = "3872051d7a6b80c2b69ceb5e4db649cb"

DB_PATH = "banco_rh.db"

def formatar_setor(setor_nome):
    if not setor_nome or setor_nome == "Outros / Não Informado" or setor_nome == "-": return "Outros"
    s = str(setor_nome).strip().title()
    replaces = {"Rh": "RH", "Dho": "DHO", "Ti": "TI", "Pcp": "PCP", "Uchoa": "Uchoa"}
    return replaces.get(s, s)

def iniciar_banco_dados():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''CREATE TABLE IF NOT EXISTS notion_cache (tabela_nome TEXT PRIMARY KEY, dados_json TEXT, ultima_atualizacao TIMESTAMP)''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS historico_folha (id INTEGER PRIMARY KEY AUTOINCREMENT, mes INTEGER, ano INTEGER, nome_funcionario TEXT, setor TEXT, salario_base REAL, horas_desconto TEXT, valor_desconto REAL, data_lancamento TIMESTAMP, faltas_dias REAL DEFAULT 0.0)''')
    try: cursor.execute("ALTER TABLE historico_folha ADD COLUMN faltas_dias REAL DEFAULT 0.0")
    except: pass
    conn.commit()
    conn.close()

iniciar_banco_dados()

def salvar_no_banco(tabela_nome, dados_lista):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    json_str = json.dumps(dados_lista)
    agora = datetime.now()
    cursor.execute('''INSERT INTO notion_cache (tabela_nome, dados_json, ultima_atualizacao) VALUES (?, ?, ?) ON CONFLICT(tabela_nome) DO UPDATE SET dados_json = excluded.dados_json, ultima_atualizacao = excluded.ultima_atualizacao''', (tabela_nome, json_str, agora))
    conn.commit()
    conn.close()

def ler_do_banco(tabela_nome):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT dados_json FROM notion_cache WHERE tabela_nome = ?", (tabela_nome,))
    resultado = cursor.fetchone()
    conn.close()
    if resultado: return json.loads(resultado[0])
    return None

class LoginData(BaseModel):
    usuario: str
    senha: str

USUARIOS_PERMITIDOS = {"diretoria": "@senha123", "gerencia": "@senha456", "rh": "@rh2026"}

@app.post("/api/login")
def validar_login(dados: LoginData):
    usuario_digitado = dados.usuario.lower()
    if usuario_digitado in USUARIOS_PERMITIDOS and USUARIOS_PERMITIDOS[usuario_digitado] == dados.senha: return {"sucesso": True, "usuario": usuario_digitado}
    return {"sucesso": False, "mensagem": "Usuário ou senha incorretos."}

def buscar_itens_notion(database_id, payload_filtro=None):
    if not database_id or database_id.startswith("ID_"): return []
    headers = {"Authorization": f"Bearer {NOTION_TOKEN}", "Notion-Version": "2022-06-28", "Content-Type": "application/json"}
    url = f"https://api.notion.com/v1/databases/{database_id}/query"
    itens, tem_mais_paginas, next_cursor = [], True, None
    if payload_filtro is None: payload_filtro = {}
    while tem_mais_paginas:
        if next_cursor: payload_filtro["start_cursor"] = next_cursor
        try:
            resp = requests.post(url, headers=headers, json=payload_filtro, verify=False, timeout=30)
            if resp.status_code == 200:
                dados = resp.json()
                itens.extend(dados["results"])
                tem_mais_paginas = dados.get("has_more")
                next_cursor = dados.get("next_cursor")
            elif resp.status_code == 429: 
                time.sleep(1.5) 
                continue
            else: break
        except Exception: break
    return itens

def extrair_texto(propriedades, nome_coluna):
    try:
        if nome_coluna in propriedades:
            prop = propriedades[nome_coluna]
            if prop["type"] == "select" and prop.get("select"): return prop["select"]["name"]
            if prop["type"] == "rich_text" and prop.get("rich_text"): return prop["rich_text"][0]["plain_text"]
            if prop["type"] == "rollup" and prop.get("rollup"):
                arr = prop["rollup"].get("array", [])
                if arr and arr[0].get("title"): return arr[0]["title"][0]["plain_text"]
                if arr and arr[0].get("rich_text"): return arr[0]["rich_text"][0]["plain_text"]
                if arr and arr[0].get("select"): return arr[0]["select"]["name"]
            if prop["type"] == "title" and prop.get("title"): return prop["title"][0]["plain_text"]
            if prop["type"] == "number" and prop.get("number") is not None: return str(prop["number"])
            if prop["type"] == "date" and prop.get("date"): return prop["date"]["start"]
    except: pass
    return "Outros / Não Informado"

def normalizar_nome(nome):
    if not nome or nome == "Outros / Não Informado": 
        return ""
    
    n = str(nome).strip().upper()
    # 1. Remove os acentos
    n = unicodedata.normalize('NFKD', n).encode('ASCII', 'ignore').decode('utf-8')
    
    # 2. Filtro inteligente para ignorar preposições e erros de digitação comuns
    palavras = n.split()
    palavras_limpas = [p for p in palavras if p not in ["DE", "DA", "DO", "DAS", "DOS", "E"]]
    
    # 3. Junta tudo de novo com espaços perfeitos
    return " ".join(palavras_limpas)

def sincronizar_tudo():
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        f1 = executor.submit(buscar_itens_notion, DB_COLABORADORES)
        f2 = executor.submit(buscar_itens_notion, DB_DESLIGAMENTOS)
        f3 = executor.submit(buscar_itens_notion, DB_ATESTADOS)
        f4 = executor.submit(buscar_itens_notion, DB_ADVERTENCIAS)
        f5 = executor.submit(buscar_itens_notion, DB_FREQUENCIA)
        f6 = executor.submit(buscar_itens_notion, DB_DESEMPENHO)
        f7 = executor.submit(buscar_itens_notion, DB_ARMARIOS)
        salvar_no_banco("colaboradores", f1.result())
        salvar_no_banco("desligamentos", f2.result())
        salvar_no_banco("atestados", f3.result())
        salvar_no_banco("advertencias", f4.result())
        salvar_no_banco("frequencia", f5.result())
        salvar_no_banco("desempenho", f6.result())
        salvar_no_banco("armarios", f7.result())

@app.get("/api/sincronizar")
def endpoint_sincronizar(background_tasks: BackgroundTasks):
    background_tasks.add_task(sincronizar_tudo)
    return {"sucesso": True, "mensagem": "Sincronização iniciada."}

@app.post("/api/processar_ponto")
async def processar_arquivo_ponto(arquivo: UploadFile = File(...)):
    try:
        conteudo = await arquivo.read()
        try: df = pd.read_excel(io.BytesIO(conteudo), header=None, engine='xlrd')
        except Exception as e: return {"sucesso": False, "erro": f"Erro Técnico do Python: {str(e)}"}
        colabs = ler_do_banco("colaboradores") or []
        dict_salarios = {}
        for c in colabs:
            props = c.get("properties", {})
            nome = extrair_texto(props, "Funcionário")
            if nome == "Outros / Não Informado": nome = extrair_texto(props, "Nome")
            if nome == "Outros / Não Informado":
                for k, v in props.items():
                    if v.get("type") == "title" and v.get("title"):
                        nome = v["title"][0]["plain_text"]
                        break
            sal_str = extrair_texto(props, "Salário (R$)")
            salario = 0.0
            try: salario = float(sal_str)
            except: pass
            setor = extrair_texto(props, "Setor")
            nome_norm = normalizar_nome(nome)
            dict_salarios[nome_norm] = {"salario": salario, "setor": setor, "nome_original": nome}

        resultados = []
        nome_atual_sap = None
        setor_atual_sap = None
        faltas_dias_atual = 0.0
        
        # LISTA NEGRA TEMPORÁRIA
        nomes_ignorados = [
            "Ana Carolina Sant Ana Vieira", "Debora Perpetua Barbosa", "Elias Honório Garcia",
            "Gabriel Felipe Aparecido De Moraes Braz", "Gleice Kely Da Silva Rodrigues Barroso",
            "Priscila Aparecida Cerqueira", "Vitoria Carolina Novaes Luiz"
        ]
        
        for i, row in df.iterrows():
            row_str = ' | '.join([str(x) if pd.notna(x) else "" for x in row.values])
            if 'Funcionário' in row_str and ':' in row_str:
                parts = row_str.split('Funcionário')[1].split(':')
                if len(parts) > 1:
                    emp_info = parts[1].replace('|', '').strip()
                    emp_parts = emp_info.split(' ', 1)
                    if len(emp_parts) == 2:
                        nome_temp = normalizar_nome(emp_parts[1])
                        pular_funcionario = False
                        for n_block in nomes_ignorados:
                            if normalizar_nome(n_block) in nome_temp:
                                pular_funcionario = True
                                break
                        if pular_funcionario: nome_atual_sap = None
                        else: nome_atual_sap = nome_temp
                        setor_atual_sap = None 
                        faltas_dias_atual = 0.0
                        
            if 'Setor' in row_str and ':' in row_str:
                parts = row_str.split(':')
                if len(parts) > 1:
                    sector_name_parts = parts[1].split('|')
                    found = [s.strip() for s in sector_name_parts if s.strip() and not s.strip().isdigit()]
                    if found: setor_atual_sap = formatar_setor(found[0])
            if setor_atual_sap and ("UCHOA" in setor_atual_sap.upper() or "UCHÔA" in setor_atual_sap.upper()):
                if 'Tot Descontado' in row_str: nome_atual_sap = None 
                continue 

            if nome_atual_sap and 'FALTA' in row_str:
                vals = []
                for j in [4, 5, 6, 7]: 
                    if j < len(row.values) and pd.notna(row.values[j]) and str(row.values[j]).strip() != "":
                        vals.append(str(row.values[j]).strip().upper())
                fc = sum(1 for v in vals if 'FALTA' in v)
                if fc == 4 or fc == 3: faltas_dias_atual += 1.0   
                elif fc == 2: faltas_dias_atual += 0.5   
                elif fc == 1: faltas_dias_atual += 0.0   
            
            if 'Tot Descontado' in row_str and nome_atual_sap:
                time_val = None
                for cell in row.values:
                    if pd.notna(cell) and isinstance(cell, str) and ':' in cell and len(cell.strip()) <= 6:
                        time_val = cell.strip()
                        break
                if time_val:
                    try:
                        info_func = dict_salarios.get(nome_atual_sap)
                        if info_func:
                            h, m = time_val.split(":")
                            horas_decimais = int(h) + (int(m) / 60.0)
                            salario = info_func["salario"]
                            if salario <= 0: salario = 2270.22 
                            valor_hora = salario / 220.0
                            desconto_rs = horas_decimais * valor_hora
                            resultados.append({
                                "nome": info_func["nome_original"],
                                "setor": formatar_setor(info_func.get("setor")),
                                "salario_base": salario,
                                "horas_desconto": time_val,
                                "valor_desconto": round(desconto_rs, 2),
                                "faltas_dias": faltas_dias_atual
                            })
                        else:
                            nome_bonito = emp_parts[1].strip().title()
                            resultados.append({
                                "nome": f"{nome_bonito} (Não achou)",
                                "setor": setor_atual_sap if setor_atual_sap else "Outros",
                                "salario_base": 0.0,
                                "horas_desconto": time_val,
                                "valor_desconto": 0.0,
                                "faltas_dias": faltas_dias_atual
                            })
                    except: pass
                    nome_atual_sap = None 

        return {"sucesso": True, "processados": len(resultados), "dados": resultados}
    except Exception as e: return {"sucesso": False, "erro": str(e)}


    # ==========================================
# FUNÇÃO AUXILIAR PARA SOMAR TEMPOS (HH:MM)
# ==========================================
def somar_horas(h1, h2):
    if not h1: return h2
    if not h2: return h1
    try:
        def para_minutos(h_str):
            partes = h_str.split(':')
            return int(partes[0]) * 60 + int(partes[1])
        
        total_min = para_minutos(h1) + para_minutos(h2)
        horas = total_min // 60
        minutos = total_min % 60
        return f"{horas:02d}:{minutos:02d}"
    except:
        return h1

# ==========================================
# MOTOR RPA 3.0 - REGRA ABSOLUTA RH (FALTAS E DSR)
# ==========================================
@app.post("/api/rpa_horas_extras")
async def rpa_horas_extras(arquivo_sap: UploadFile = File(...), arquivo_escritorio: UploadFile = File(...)):
    try:
        conteudo_sap = await arquivo_sap.read()
        df_sap = pd.read_excel(io.BytesIO(conteudo_sap), header=None, engine='xlrd')

        dados_sap = {}
        current_norm = None
        faltas_dias_atual = 0.0
        dsr_perdidos_atual = 0
        teve_falta_integral_na_semana = False

        for i, row in df_sap.iterrows():
            row_str = ' | '.join([str(x).strip() if pd.notna(x) else "" for x in row.values])
            
            # Localiza o funcionário
            if 'Funcionário' in row_str and ':' in row_str:
                # Salva o anterior antes de limpar as variáveis
                if current_norm and current_norm in dados_sap:
                    # Se o mês cortou no meio da semana e ficou uma falta pendente, cobra o DSR
                    if teve_falta_integral_na_semana: 
                        dsr_perdidos_atual += 1
                    
                    dados_sap[current_norm]["faltas_dias"] = faltas_dias_atual
                    dados_sap[current_norm]["dsr_perdidos"] = dsr_perdidos_atual
                
                parts = row_str.split(':')
                if len(parts) > 1:
                    emp_info = parts[1].replace('|', '').strip()
                    emp_parts = emp_info.split(' ', 1)
                    if len(emp_parts) == 2:
                        current_norm = normalizar_nome(emp_parts[1])
                        dados_sap[current_norm] = {"he_50": None, "he_100": None, "adc_noturno": None, "faltas_dias": 0.0, "dsr_perdidos": 0}
                        faltas_dias_atual = 0.0
                        dsr_perdidos_atual = 0
                        teve_falta_integral_na_semana = False
                    else:
                        current_norm = None
                        
            if current_norm:
                # 1. Conta Faltas Diárias (Apenas dias inteiros somam)
                if 'FALTA' in row_str:
                    vals = []
                    for j in [4, 5, 6, 7]: 
                        if j < len(row.values) and pd.notna(row.values[j]) and str(row.values[j]).strip() != "":
                            vals.append(str(row.values[j]).strip().upper())
                    fc = sum(1 for v in vals if 'FALTA' in v)
                    
                    if fc >= 3: 
                        faltas_dias_atual += 1.0
                        teve_falta_integral_na_semana = True # Aciona o gatilho da punição do DSR

                # 2. Bateu no Domingo, fecha a semana
                if 'DOM' in row_str.upper():
                    # REGRA ABSOLUTA: Teve falta integral? Perde 1 DSR. (Ignora o que o SAP diz)
                    if teve_falta_integral_na_semana:
                        dsr_perdidos_atual += 1
                        
                    # Zera a semana para começar a contar o próximo ciclo
                    teve_falta_integral_na_semana = False

                # 3. Captura Horas Extras
                if 'Extra A 050%' in row_str:
                    match = re.search(r'Extra A 050%\s*:\s*(\d{1,3}:\d{2})', row_str)
                    if match: dados_sap[current_norm]["he_50"] = somar_horas(dados_sap[current_norm]["he_50"], match.group(1))

                if 'Ext Adi A 050%' in row_str:
                    match = re.search(r'Ext Adi A 050%\s*:\s*(\d{1,3}:\d{2})', row_str)
                    if match: dados_sap[current_norm]["he_50"] = somar_horas(dados_sap[current_norm]["he_50"], match.group(1))
                    
                if 'Extra A 100%' in row_str:
                    match = re.search(r'Extra A 100%\s*:\s*(\d{1,3}:\d{2})', row_str)
                    if match: dados_sap[current_norm]["he_100"] = somar_horas(dados_sap[current_norm]["he_100"], match.group(1))

                if 'Ext Adi A 100%' in row_str:
                    match = re.search(r'Ext Adi A 100%\s*:\s*(\d{1,3}:\d{2})', row_str)
                    if match: dados_sap[current_norm]["he_100"] = somar_horas(dados_sap[current_norm]["he_100"], match.group(1))

                if 'Adc Noturno' in row_str:
                    match = re.search(r'(\d{1,3}:\d{2})', row_str.replace('Adc Noturno', ''))
                    if match: dados_sap[current_norm]["adc_noturno"] = match.group(1)

        # Salva o último funcionário do loop
        if current_norm and current_norm in dados_sap:
            if teve_falta_integral_na_semana: dsr_perdidos_atual += 1
            dados_sap[current_norm]["faltas_dias"] = faltas_dias_atual
            dados_sap[current_norm]["dsr_perdidos"] = dsr_perdidos_atual

        # 2. Injeta no Escritório
        conteudo_escritorio = await arquivo_escritorio.read()
        wb = openpyxl.load_workbook(io.BytesIO(conteudo_escritorio))

        for ws in wb.worksheets:
            col_nome = col_he50 = col_he100 = col_noturno = col_faltas_dsr = header_row = None
            
            for r in range(1, 15):
                for c in range(1, ws.max_column + 1):
                    val = str(ws.cell(row=r, column=c).value or "").strip().lower()
                    if 'nome' in val and 'colaborador' in val: col_nome = c
                    elif 'adc 50%' in val: col_he50 = c
                    elif 'adc 100%' in val: col_he100 = c
                    elif 'noturno' in val: col_noturno = c
                    elif 'faltas' in val and 'dsr' in val: col_faltas_dsr = c
                if col_nome and col_he50:
                    header_row = r
                    break

            if col_nome and col_he50 and header_row:
                for r in range(header_row + 1, ws.max_row + 1):
                    nome_cell = ws.cell(row=r, column=col_nome).value
                    if nome_cell:
                        norm_excel = normalizar_nome(str(nome_cell))
                        if norm_excel in dados_sap:
                            info = dados_sap[norm_excel]
                            
                            if info["he_50"] and col_he50: ws.cell(row=r, column=col_he50).value = info["he_50"]
                            if info["he_100"] and col_he100: ws.cell(row=r, column=col_he100).value = info["he_100"]
                            if info.get("adc_noturno") and col_noturno: ws.cell(row=r, column=col_noturno).value = info["adc_noturno"]
                                
                            # Preenche Faltas / DSR
                            if info.get("faltas_dias", 0) > 0 and col_faltas_dsr:
                                qtd_f = info["faltas_dias"]
                                qtd_dsr = info.get("dsr_perdidos", 0)
                                
                                f_str = str(int(qtd_f)) if qtd_f.is_integer() else str(qtd_f)
                                d_str = str(int(qtd_dsr))
                                
                                ws.cell(row=r, column=col_faltas_dsr).value = f"{f_str}+{d_str}dsr"

        saida_memoria = io.BytesIO()
        wb.save(saida_memoria)
        saida_memoria.seek(0)

        return StreamingResponse(
            saida_memoria,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=ESCRITORIO_PRONTO.xlsx", "Access-Control-Expose-Headers": "Content-Disposition"}
        )
    except Exception as e:
        return {"sucesso": False, "erro": str(e)}

class DadosFolha(BaseModel):
    mes: int
    ano: int
    lancamentos: list

@app.post("/api/salvar_folha")
def salvar_dados_folha(dados: DadosFolha):
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        agora = datetime.now()
        cursor.execute("DELETE FROM historico_folha WHERE mes = ? AND ano = ?", (dados.mes, dados.ano))
        for lanc in dados.lancamentos:
            cursor.execute('''INSERT INTO historico_folha (mes, ano, nome_funcionario, setor, salario_base, horas_desconto, valor_desconto, data_lancamento, faltas_dias) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''', (dados.mes, dados.ano, lanc['nome'], lanc['setor'], lanc['salario_base'], lanc['horas_desconto'], lanc['valor_desconto'], agora, lanc.get('faltas_dias', 0.0)))
        conn.commit()
        conn.close()
        return {"sucesso": True, "mensagem": f"{len(dados.lancamentos)} lançamentos salvos com sucesso no Banco de Dados!"}
    except Exception as e: return {"sucesso": False, "erro": str(e)}

@app.get("/api/dashboard/kpis")
def obter_kpis_do_banco(mes: int = None, ano: int = None, setor: str = "Todos"):
    hoje = datetime.now()
    mes_int = mes if mes else hoje.month
    ano_int = ano if ano else hoje.year

    inicio_mes_civil = f"{ano_int}-{mes_int:02d}-01"
    ultimo_dia = calendar.monthrange(ano_int, mes_int)[1]
    fim_mes_civil = f"{ano_int}-{mes_int:02d}-{ultimo_dia}"

    if mes_int == 1: mes_anterior, ano_anterior = 12, ano_int - 1
    else: mes_anterior, ano_anterior = mes_int - 1, ano_int
    inicio_mes_fiscal = f"{ano_anterior}-{mes_anterior:02d}-26"
    fim_mes_fiscal = f"{ano_int}-{mes_int:02d}-25"

    todos_colab = ler_do_banco("colaboradores") or []
    todos_deslig = ler_do_banco("desligamentos") or []
    todos_atestados = ler_do_banco("atestados") or []
    todos_adv = ler_do_banco("advertencias") or []
    todos_freq = ler_do_banco("frequencia") or []
    avaliacoes_itens = ler_do_banco("desempenho") or []
    armarios_itens = ler_do_banco("armarios") or []

    ativos_itens = [i for i in todos_colab if extrair_texto(i.get("properties", {}), "Status") == "Ativo"]
    admissoes_itens = [i for i in todos_colab if inicio_mes_civil <= extrair_texto(i.get("properties", {}), "Data de admissão")[:10] <= fim_mes_civil]
    
    avaliacoes_filtradas = []
    for item in avaliacoes_itens:
        props = item.get("properties", {})
        dt_aval = extrair_texto(props, "Data da Avaliação")
        if dt_aval == "Outros / Não Informado": dt_aval = extrair_texto(props, "Data")
        if dt_aval != "Outros / Não Informado" and inicio_mes_civil <= dt_aval[:10] <= fim_mes_civil: avaliacoes_filtradas.append(item)
    avaliacoes_itens = avaliacoes_filtradas
    
    desligamentos_itens = [i for i in todos_deslig if inicio_mes_fiscal <= extrair_texto(i.get("properties", {}), "Data de Desligamento")[:10] <= fim_mes_fiscal]
    desligamentos_ano = [i for i in todos_deslig if extrair_texto(i.get("properties", {}), "Data de Desligamento")[:4] == str(ano_int)]
    atestados_itens = [i for i in todos_atestados if inicio_mes_fiscal <= extrair_texto(i.get("properties", {}), "Data de Entrega")[:10] <= fim_mes_fiscal]
    advertencias_itens = [i for i in todos_adv if inicio_mes_fiscal <= extrair_texto(i.get("properties", {}), "Data da Advertência")[:10] <= fim_mes_fiscal]

    setores_unicos = set([formatar_setor(extrair_texto(i.get("properties", {}), "Setor")) for i in ativos_itens])
    if "Uchoa" in setores_unicos: setores_unicos.remove("Uchoa")
    lista_setores = sorted(list(setores_unicos))

    if setor != "Todos":
        ativos_itens = [i for i in ativos_itens if formatar_setor(extrair_texto(i.get("properties", {}), "Setor")) == setor]
        admissoes_itens = [i for i in admissoes_itens if formatar_setor(extrair_texto(i.get("properties", {}), "Setor")) == setor]
        desligamentos_itens = [i for i in desligamentos_itens if formatar_setor(extrair_texto(i.get("properties", {}), "Setor")) == setor]
        atestados_itens = [i for i in atestados_itens if formatar_setor(extrair_texto(i.get("properties", {}), "Setor")) == setor]
        advertencias_itens = [i for i in advertencias_itens if formatar_setor(extrair_texto(i.get("properties", {}), "Setor")) == setor]
        desligamentos_ano = [i for i in desligamentos_ano if formatar_setor(extrair_texto(i.get("properties", {}), "Setor")) == setor]
        avaliacoes_itens = [i for i in avaliacoes_itens if formatar_setor(extrair_texto(i.get("properties", {}), "Setor")) == setor]
        armarios_itens = [i for i in armarios_itens if formatar_setor(extrair_texto(i.get("properties", {}), "Setor")) == setor]

    total_ativos = len(ativos_itens)
    dict_perfis = {}
    
    def get_nome_correto(props, is_atestado=False):
        nome = "Outros / Não Informado"
        if "Funcionário" in props:
            prop = props["Funcionário"]
            if prop["type"] == "rich_text" and prop.get("rich_text"): nome = prop["rich_text"][0]["plain_text"]
            elif prop["type"] == "title" and prop.get("title"): nome = prop["title"][0]["plain_text"]
            elif prop["type"] == "rollup" and prop.get("rollup"): 
                arr = prop["rollup"].get("array", [])
                if arr and arr[0].get("title"): nome = arr[0]["title"][0]["plain_text"]
                elif arr and arr[0].get("rich_text"): nome = arr[0]["rich_text"][0]["plain_text"]
        
        if nome == "Outros / Não Informado" and "Nome" in props: nome = extrair_texto(props, "Nome")
        
        if is_atestado and (nome == "Outros / Não Informado" or not nome.strip()):
            for k, v in props.items():
                if v.get("type") == "title" and v.get("title"):
                    nome_bruto = v["title"][0]["plain_text"]
                    nome = re.split(r'[-–—]', nome_bruto)[0].strip()
                    break
                    
        return nome if nome else "Outros / Não Informado"

    def iniciar_perfil(nome):
        if nome not in dict_perfis:
            dict_perfis[nome] = {"cargo": "-", "setor": "-", "tempo_casa": "-", "salario": 0.0, "historico_atestados": [], "historico_advertencias": [], "faltas_dias": 0.0, "atrasos": 0, "nota_desempenho": 0.0, "qtd_aval": 0}

    alertas_aniversarios, alertas_contratos, alertas_ferias = [], [], []
    for item in ativos_itens:
        props = item.get("properties", {})
        nome = get_nome_correto(props)
        iniciar_perfil(nome)
        dict_perfis[nome]["setor"] = formatar_setor(extrair_texto(props, "Setor"))

        adm_str = extrair_texto(props, "Data de admissão")
        if adm_str != "Outros / Não Informado":
            try:
                data_admissao = datetime.strptime(adm_str[:10], "%Y-%m-%d")
                diff = hoje - data_admissao
                anos, meses = diff.days // 365, (diff.days % 365) // 30
                dict_perfis[nome]["tempo_casa"] = f"{anos} ano(s) e {meses} mês(es)" if anos > 0 else f"{meses} mês(es)"
                if diff.days >= 365:
                    dias_para_dobrar = (365 * 2) - diff.days
                    if -365 < dias_para_dobrar <= 120: alertas_ferias.append({"nome": nome, "dias_restantes": dias_para_dobrar, "setor": dict_perfis[nome]["setor"]})
                venc_45, venc_90 = data_admissao + timedelta(days=45), data_admissao + timedelta(days=90)
                if venc_45.year == ano_int and venc_45.month == mes_int: alertas_contratos.append({"nome": nome, "dia": venc_45.day, "tipo": "45 Dias"})
                if venc_90.year == ano_int and venc_90.month == mes_int: alertas_contratos.append({"nome": nome, "dia": venc_90.day, "tipo": "90 Dias"})
            except: pass
            
        nasc_str = extrair_texto(props, "Data de Nascimento")
        if nasc_str == "Outros / Não Informado": nasc_str = extrair_texto(props, "Nascimento")
        if nasc_str != "Outros / Não Informado":
            try:
                if int(nasc_str.split("-")[1]) == mes_int: alertas_aniversarios.append({"nome": nome, "dia": int(nasc_str.split("-")[2])})
            except: pass

    alertas_aniversarios = sorted(alertas_aniversarios, key=lambda x: x["dia"])
    alertas_contratos = sorted(alertas_contratos, key=lambda x: x["dia"])
    alertas_ferias = sorted(alertas_ferias, key=lambda x: x["dias_restantes"])

    dict_setores = {}
    for item in atestados_itens:
        s = formatar_setor(extrair_texto(item.get("properties", {}), "Setor"))
        if "Uchoa" in s: continue
        if s not in dict_setores: dict_setores[s] = {"setor": s, "atestados": 0, "faltas": 0}
        dict_setores[s]["atestados"] += 1

    total_faltas_inteiras, total_atrasos, total_horas_extras = 0.0, 0, 0.0
    dict_ranking, dict_he_setor = {}, {}
    total_perdas_r = 0.0

    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        if setor == "Todos": cursor.execute("SELECT nome_funcionario, setor, valor_desconto, faltas_dias FROM historico_folha WHERE mes = ? AND ano = ?", (mes_int, ano_int))
        else: cursor.execute("SELECT nome_funcionario, setor, valor_desconto, faltas_dias FROM historico_folha WHERE mes = ? AND ano = ? COLLATE NOCASE", (mes_int, ano_int))
        for rec in cursor.fetchall():
            setor_f = formatar_setor(rec[1])
            if "Uchoa" in setor_f: continue
            if setor != "Todos" and setor_f != formatar_setor(setor): continue
            nome_f = rec[0]
            val_desc = rec[2] if rec[2] is not None else 0.0
            faltas_d = rec[3] if rec[3] is not None else 0.0
            total_perdas_r += val_desc
            total_faltas_inteiras += faltas_d
            iniciar_perfil(nome_f)
            dict_perfis[nome_f]["faltas_dias"] += faltas_d
            dict_ranking[nome_f] = dict_ranking.get(nome_f, 0) + faltas_d
            if setor_f not in dict_setores: dict_setores[setor_f] = {"setor": setor_f, "atestados": 0, "faltas": 0}
            dict_setores[setor_f]["faltas"] += faltas_d
        conn.close()
    except Exception as e: pass

    for item in todos_freq:
        dt = extrair_texto(item.get("properties", {}), "Data")
        if dt != "Outros / Não Informado" and inicio_mes_fiscal <= dt[:10] <= fim_mes_fiscal:
            props = item.get("properties", {})
            setor_freq = formatar_setor(extrair_texto(props, "Setor"))
            if "Uchoa" in setor_freq: continue
            
            if setor != "Todos" and setor_freq != formatar_setor(setor): continue
            
            nome = get_nome_correto(props)
            iniciar_perfil(nome)
            prop_dias = props.get("Dias") or props.get("# Dias") or {}
            dias_descontados = prop_dias.get("number") if prop_dias.get("type") == "number" else 0
            if not dias_descontados or dias_descontados == 0:
                total_atrasos += 1
                dict_perfis[nome]["atrasos"] += 1
            prop_he = props.get("Horas Extras") or props.get("HE") or props.get("Valor HE") or {}
            qtd_he = prop_he.get("number") if prop_he.get("type") == "number" else 0
            if qtd_he and qtd_he > 0:
                total_horas_extras += qtd_he
                if setor_freq not in dict_he_setor: dict_he_setor[setor_freq] = 0
                dict_he_setor[setor_freq] += qtd_he

    ranking_faltas = sorted([{"nome": k, "faltas": v} for k, v in dict_ranking.items()], key=lambda x: x["faltas"], reverse=True)[:10]
    grafico_he = sorted([{"setor": k, "horas": v} for k, v in dict_he_setor.items()], key=lambda x: x["horas"], reverse=True)
    grafico_setores = list(dict_setores.values())

    dict_ranking_atestados, dict_medicos, dict_cids = {}, {}, {}
    for item in atestados_itens:
        props = item.get("properties", {})
        setor_atst = formatar_setor(extrair_texto(props, "Setor"))
        if "Uchoa" in setor_atst: continue
        
        nome = get_nome_correto(props, is_atestado=True)
        
        iniciar_perfil(nome)
        dict_ranking_atestados[nome] = dict_ranking_atestados.get(nome, 0) + 1
        medico = extrair_texto(props, "Médico")
        if medico and medico != "Outros / Não Informado": dict_medicos[medico] = dict_medicos.get(medico, 0) + 1
        cid = extrair_texto(props, "CID")
        motivo = extrair_texto(props, "Motivo")
        data_str = extrair_texto(props, "Data de Entrega")
        if cid and cid != "Outros / Não Informado":
            label_cid = f"{cid} - {motivo}" if motivo != "Outros / Não Informado" else cid
            dict_cids[label_cid] = dict_cids.get(label_cid, 0) + 1
            dict_perfis[nome]["historico_atestados"].append({"data": data_str[:10] if data_str != "Outros / Não Informado" else "-", "motivo": label_cid})
        else:
            dict_perfis[nome]["historico_atestados"].append({"data": data_str[:10] if data_str != "Outros / Não Informado" else "-", "motivo": motivo})
    
    ranking_atestados = sorted([{"nome": k, "atestados": v} for k, v in dict_ranking_atestados.items()], key=lambda x: x["atestados"], reverse=True)[:10]
    ranking_medicos = sorted([{"nome": k, "quantidade": v} for k, v in dict_medicos.items()], key=lambda x: x["quantidade"], reverse=True)[:7]
    ranking_cids = sorted([{"nome": k, "quantidade": v} for k, v in dict_cids.items()], key=lambda x: x["quantidade"], reverse=True)[:10]

    dict_adv, dict_ranking_adv = {}, {}
    for item in advertencias_itens:
        props = item.get("properties", {})
        setor_adv = formatar_setor(extrair_texto(props, "Setor"))
        if "Uchoa" in setor_adv: continue
        nome = get_nome_correto(props)
        iniciar_perfil(nome)
        dict_ranking_adv[nome] = dict_ranking_adv.get(nome, 0) + 1
        motivo = extrair_texto(props, "Motivo")
        if motivo == "Outros / Não Informado": motivo = extrair_texto(props, "Tipo")
        dict_adv[motivo] = dict_adv.get(motivo, 0) + 1
        data_str = extrair_texto(props, "Data da Advertência")
        dict_perfis[nome]["historico_advertencias"].append({"data": data_str[:10] if data_str != "Outros / Não Informado" else "-", "motivo": motivo})
        
    grafico_advertencias = [{"name": k, "value": v} for k, v in dict_adv.items()]
    ranking_advertencias = sorted([{"nome": k, "advertencias": v} for k, v in dict_ranking_adv.items()], key=lambda x: x["advertencias"], reverse=True)[:10]

    competencias = ["Comunicação", "Produtividade", "Trabalho em Equipe", "Proatividade", "Pontualidade"]
    radar_somas, radar_cont = {c: 0 for c in competencias}, {c: 0 for c in competencias}
    for item in avaliacoes_itens:
        props = item.get("properties", {})
        nome = get_nome_correto(props)
        iniciar_perfil(nome)
        soma_func, qtd_func = 0, 0
        for c in competencias:
            prop = props.get(c, {})
            if prop.get("type") == "number" and prop.get("number") is not None:
                radar_somas[c] += prop.get("number")
                radar_cont[c] += 1
                soma_func += prop.get("number")
                qtd_func += 1
        if qtd_func > 0:
            media_func = soma_func / qtd_func
            curr_nota = dict_perfis[nome]["nota_desempenho"]
            curr_qtd = dict_perfis[nome]["qtd_aval"]
            dict_perfis[nome]["nota_desempenho"] = ((curr_nota * curr_qtd) + media_func) / (curr_qtd + 1)
            dict_perfis[nome]["qtd_aval"] += 1
            
    grafico_radar = [{"subject": c, "A": round(radar_somas[c] / radar_cont[c], 1) if radar_cont[c] > 0 else 0, "fullMark": 5} for c in competencias]

    meses_nomes = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    contagem_meses, dict_motivos = {m: 0 for m in meses_nomes}, {}
    for item in desligamentos_ano:
        props = item.get("properties", {})
        dt = extrair_texto(props, "Data de Desligamento")
        if dt != "Outros / Não Informado":
            mes_idx = int(dt.split("-")[1]) - 1
            contagem_meses[meses_nomes[mes_idx]] += 1
        motivo = extrair_texto(props, "Motivo de Desligamento")
        if motivo not in dict_motivos: dict_motivos[motivo] = 0
        dict_motivos[motivo] += 1
        
    grafico_turnover = [{"mes": mes, "turnover": round((contagem_meses[mes] / total_ativos) * 100, 1) if total_ativos > 0 else 0} for mes in meses_nomes]
    grafico_motivos = [{"name": k, "value": v} for k, v in dict_motivos.items()]

    dict_headcount = {}
    for item in ativos_itens:
        s = formatar_setor(extrair_texto(item.get("properties", {}), "Setor"))
        if "Uchoa" in s: continue
        if s not in dict_headcount: dict_headcount[s] = 0
        dict_headcount[s] += 1
    grafico_headcount = sorted([{"setor": k, "quantidade": v} for k, v in dict_headcount.items()], key=lambda x: x["quantidade"], reverse=True)
    turnover_atual = f"{(len(desligamentos_itens) / total_ativos) * 100:.1f}%" if total_ativos > 0 else "0.0%"

    lista_armarios = []
    for item in armarios_itens:
        props = item.get("properties", {})
        num_str = extrair_texto(props, "armario_numero")
        if num_str == "Outros / Não Informado":
             for k, v in props.items():
                if v.get("type") == "title" and v.get("title"):
                    num_str = v["title"][0]["plain_text"]
                    break
        num_val = 0
        try:
             clean_num = ''.join(filter(str.isdigit, num_str))
             num_val = int(clean_num) if clean_num else 0
        except: pass

        dono = extrair_texto(props, "nome_funcionario")
        if dono == "Outros / Não Informado":
            for k, v in props.items():
                if v.get("type") == "title" and v.get("title"):
                    dono = v["title"][0]["plain_text"]
                    break
        if dono == "Outros / Não Informado": dono = None

        status = extrair_texto(props, "status")
        status_clean = "Livre"
        if not status or status == "Outros / Não Informado":
             if dono: status_clean = "Ocupado"
             else: status_clean = "Livre"
        else:
            status_lower = status.lower()
            if "ocupado" in status_lower: status_clean = "Ocupado"
            elif "manuten" in status_lower or "indispon" in status_lower: status_clean = "Manutenção"
            elif "livre" in status_lower or "dispon" in status_lower: status_clean = "Livre"
            
        lista_armarios.append({"num": num_str if num_str != "Outros / Não Informado" else "?", "sort_val": num_val, "dono": dono, "status": status_clean})

    lista_armarios = sorted(lista_armarios, key=lambda x: x["sort_val"])

    return {
        "funcionarios": total_ativos, "admissoes": len(admissoes_itens), "desligamentos": len(desligamentos_itens), "turnover": turnover_atual,     
        "atestados": len(atestados_itens), "advertencias": len(advertencias_itens), 
        "faltas": total_faltas_inteiras, "atrasos": total_atrasos,
        "custo_absenteismo": total_perdas_r, "avaliacoes": len(avaliacoes_itens),
        "graficoSetores": grafico_setores, "graficoTurnover": grafico_turnover, "graficoMotivos": grafico_motivos, "graficoHeadcount": grafico_headcount,
        "alertasAniversarios": alertas_aniversarios, "alertasContratos": alertas_contratos,
        "graficoAdvertencias": grafico_advertencias, 
        "rankingFaltas": ranking_faltas, 
        "rankingAtestados": ranking_atestados, 
        "rankingAdvertencias": ranking_advertencias,
        "rankingMedicos": ranking_medicos, "rankingCids": ranking_cids,       
        "graficoRadar": grafico_radar, "perfis360": dict_perfis,
        "alertasFerias": alertas_ferias, "totalHorasExtras": total_horas_extras, "graficoHorasExtras": grafico_he,
        "armarios": lista_armarios, "setoresDisponiveis": lista_setores
    }